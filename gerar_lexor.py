#!/usr/bin/env python3
"""
Converte Controle_LEXOR.xlsx (abas SIOPLEx e Ações) no módulo src/data/lexor.js
usado pela aba LEXOR do aplicativo.

Correções aplicadas na conversão (ver análise de qualidade de dados):
  1. Ação em notação científica (2E+74 -> 2E74), recuperada da Funcional Programática.
  2. Zeros à esquerda em Função (2 díg.), Subfunção (3), Programa (4).
  3. UO obtida da Funcional Programática (não da tabela Ações), que é ambígua
     quando a mesma ação existe em UOs diferentes.
  4. Objeto/Justificativa normalizados (espaços, quebras de linha, CAIXA ALTA).
  5. Tipo de emenda padronizado (27 variações -> 3 categorias + marca de repetida).
"""
import json
import re
import unicodedata
import openpyxl

ORIGEM = "/home/claude/app/Controle_LEXOR.xlsx"
DESTINO = "/home/claude/app/src/data/lexor.js"

# Siglas que devem continuar em caixa alta ao converter texto de CAIXA ALTA
SIGLAS = {
    "EB", "ESA", "OM", "TI", "GND", "UO", "CIA", "BC", "BI", "BIB", "RCB", "RCC",
    "SISFRON", "ASTROS", "IMBEL", "EASA", "CMS", "CML", "CMNE", "CMO", "CMA", "CMP",
    "CMSE", "CMN", "COTER", "DECEx", "DEC", "COLOG", "SEF", "DGP", "EME", "GSI",
    "PDF", "GPS", "UTI", "AMAN", "EsPCEx", "IME", "CMB", "PQRMNT", "MEM", "CIGE",
    "SIOP", "CNPJ", "PLOA", "LOA", "LDO", "PPA", "RP", "II", "III", "IV", "V",
    "VI", "VII", "VIII", "IX", "X", "XI", "XII", "1º", "2º", "3º", "4º", "5º",
}
PALAVRAS_MINUSCULAS = {
    "de", "da", "do", "das", "dos", "e", "em", "no", "na", "nos", "nas",
    "a", "o", "as", "os", "para", "com", "por", "ao", "aos", "à", "às", "ou",
}


def limpa(v):
    """Normaliza qualquer célula em string limpa (sem quebras de linha nem espaços duplos)."""
    if v is None:
        return ""
    s = str(v).replace("\r", " ").replace("\n", " ")
    s = re.sub(r"\s+", " ", s).strip()
    return "" if s.lower() in ("none", "nan") else s


def caixa_alta(s):
    letras = [c for c in s if c.isalpha()]
    if not letras:
        return False
    return all(c.isupper() for c in letras)


def frase(s):
    """Converte texto em CAIXA ALTA para caixa de frase, preservando siglas conhecidas."""
    s = limpa(s)
    if not s or not caixa_alta(s):
        return s[:1].upper() + s[1:] if s else s
    saida = []
    for tok in s.split(" "):
        nucleo = tok.strip("().,;:/-")
        if nucleo in SIGLAS or (len(nucleo) <= 4 and nucleo.isupper() and not any(
                c in "AEIOU" for c in nucleo)):
            saida.append(tok)
        else:
            saida.append(tok.lower())
    r = " ".join(saida)
    # primeira letra alfabética em maiúscula
    for i, c in enumerate(r):
        if c.isalpha():
            return r[:i] + c.upper() + r[i + 1:]
    return r


def num(v):
    """Converte célula em inteiro (reais), tolerando texto e vazio."""
    if v is None or v == "":
        return 0
    if isinstance(v, (int, float)):
        return int(round(v))
    s = re.sub(r"[^\d,.-]", "", str(v))
    if not s:
        return 0
    s = s.replace(".", "").replace(",", ".")
    try:
        return int(round(float(s)))
    except ValueError:
        return 0


def cientifica(v):
    """Detecta valores que o Excel corrompeu em notação científica (ex.: 2E74 -> 2e+74)."""
    if isinstance(v, float) and (abs(v) >= 1e15 or "e+" in repr(v).lower()):
        return True
    return bool(re.search(r"\de\+\d", str(v).lower()))


def partes_fp(fp):
    """Quebra a Funcional Programática '10.52921.05.363.6112.21GN.0000'."""
    p = [x.strip() for x in limpa(fp).split(".")]
    if len(p) < 7:
        return {}
    # p[0] é a esfera orçamentária (10 = Fiscal, 20 = Seguridade Social)
    return {"esfera": p[0], "uo": p[1], "funcao": p[2], "subfuncao": p[3],
            "programa": p[4], "acao": p[5].upper(), "subtitulo": p[6]}


def pad(v, n, fp_val=""):
    """Zero-padding, com fallback para o valor vindo da Funcional Programática."""
    s = limpa(v)
    if not s or cientifica(v):
        s = limpa(fp_val)
    if not s:
        return ""
    if s.replace(".", "").isdigit():
        s = s.split(".")[0]
        return s.zfill(n)
    return s


RE_REPETIDA = re.compile(r"repetid[ao]|repetido", re.I)


def classifica_tipo(t):
    """Padroniza as 27 variações da coluna Tipo."""
    t = limpa(t)
    base = unicodedata.normalize("NFKD", t.lower())
    repetida = bool(RE_REPETIDA.search(t))
    if "bancada" in base:
        cat = "Emenda de Bancada"
    elif "comiss" in base:
        cat = "Emenda de Comissão"
    elif "individual" in base:
        cat = "Emenda Individual"
    elif not t:
        cat = ""
    else:
        cat = t
    return cat, repetida


# RP conforme a LDO: 6 = individual impositiva, 7 = bancada, 8 = comissão
RP_POR_TIPO = {"Emenda Individual": "6", "Emenda de Bancada": "7", "Emenda de Comissão": "8"}


def main():
    wb = openpyxl.load_workbook(ORIGEM, data_only=True)

    # ---------- aba Ações: tabela de apoio (o "PROCV" do modelo do Word) ----------
    aba = wb["Ações"]
    acoes = {}
    uos = {}
    for r in range(2, aba.max_row + 1):
        cod = limpa(aba.cell(r, 5).value).upper()
        if not cod:
            continue
        uo_cod = limpa(aba.cell(r, 10).value)
        uo_nome = limpa(aba.cell(r, 11).value)
        if uo_cod:
            uos[uo_cod] = uo_nome
        acoes[cod] = {
            "descricao": limpa(aba.cell(r, 7).value),
            "subtitulo": limpa(aba.cell(r, 6).value) or "XXXX",
            "orgaoCod": limpa(aba.cell(r, 8).value),
            "orgaoNome": limpa(aba.cell(r, 9).value),
            "uoCod": uo_cod,
            "uoNome": uo_nome,
            "produto": limpa(aba.cell(r, 12).value),
            "unidade": limpa(aba.cell(r, 13).value) or "unidade",
            "meta": limpa(aba.cell(r, 14).value) or "1",
            "seq": limpa(aba.cell(r, 15).value),
            "cnpj": limpa(aba.cell(r, 16).value),
        }

    # ---------- aba SIOPLEx: uma proposta por linha ----------
    ws = wb["SIOPLEx"]
    props = []
    for r in range(2, ws.max_row + 1):
        nr = limpa(ws.cell(r, 1).value)
        if not nr:
            continue
        fp_bruta = ws.cell(r, 25).value
        fp = partes_fp(fp_bruta)

        acao_raw = ws.cell(r, 12).value
        acao = limpa(acao_raw).upper()
        if not acao or cientifica(acao_raw):
            acao = fp.get("acao", "")
        if acao.endswith(".0"):
            acao = acao[:-2]

        tipo, repetida = classifica_tipo(ws.cell(r, 2).value)
        uo_cod = fp.get("uo", "") or acoes.get(acao, {}).get("uoCod", "")

        # Valores originais (M/N/O) e negociados (AG/AH/AI). O espelho usa os
        # negociados; se o par negociado estiver zerado, cai nos originais.
        gnd3 = num(ws.cell(r, 13).value)
        gnd4 = num(ws.cell(r, 14).value)
        gnd3n = num(ws.cell(r, 33).value)
        gnd4n = num(ws.cell(r, 34).value)
        total = num(ws.cell(r, 15).value) or (gnd3 + gnd4)
        totaln = num(ws.cell(r, 35).value)

        registro = {
            "nr": nr,
            "tipo": tipo,
            "rep": 1 if repetida else 0,
            "proponente": limpa(ws.cell(r, 3).value),
            "ods": limpa(ws.cell(r, 4).value),
            "beneficiario": limpa(ws.cell(r, 5).value) or limpa(ws.cell(r, 20).value),
            "cidade": limpa(ws.cell(r, 6).value),
            "uf": limpa(ws.cell(r, 7).value).upper(),
            "objeto": frase(ws.cell(r, 8).value),
            "funcao": pad(ws.cell(r, 9).value, 2, fp.get("funcao")),
            "subfuncao": pad(ws.cell(r, 10).value, 3, fp.get("subfuncao")),
            "programa": pad(ws.cell(r, 11).value, 4, fp.get("programa")),
            "acao": acao,
            "uo": uo_cod,
            "esfera": fp.get("esfera", ""),
            "subtitulo": fp.get("subtitulo") or "XXXX",
            "gnd3": gnd3,
            "gnd4": gnd4,
            "gnd3n": gnd3n,
            "gnd4n": gnd4n,
            "total": total,
            "totaln": totaln,
            "justificativa": frase(ws.cell(r, 16).value),
            "cmdo": limpa(ws.cell(r, 21).value),
            "catalogo": limpa(ws.cell(r, 22).value),
            "grupo": limpa(ws.cell(r, 23).value),
            "opus": limpa(ws.cell(r, 24).value),
            "fp": limpa(fp_bruta),
            "obs": limpa(ws.cell(r, 26).value),
            "status": limpa(ws.cell(r, 29).value),
            "parlamentar": limpa(ws.cell(r, 30).value),
            "partido": limpa(ws.cell(r, 31).value).upper(),
            "rp": RP_POR_TIPO.get(tipo, "6"),
            "exportado": limpa(ws.cell(r, 40).value),
        }
        # descarta linhas em branco da planilha (só têm o Nr Proposta preenchido)
        if not registro["beneficiario"] and not registro["objeto"] and not registro["acao"]:
            continue
        # omite chaves vazias para reduzir o tamanho do pacote enviado ao navegador
        props.append({k: v for k, v in registro.items() if v not in ("", 0)})

    # ---------- relatório de consistência ----------
    sem_acao = [p["nr"] for p in props if not p.get("acao") or p.get("acao") not in acoes]
    div_uo = [p["nr"] for p in props
              if p.get("acao") in acoes and p.get("uo") and p["uo"] != acoes[p["acao"]]["uoCod"]]
    sem_valor = [p["nr"] for p in props
                 if p.get("gnd3", 0) + p.get("gnd4", 0)
                 + p.get("gnd3n", 0) + p.get("gnd4n", 0) == 0]
    negociados = [p["nr"] for p in props if p.get("gnd3n", 0) or p.get("gnd4n", 0)]
    sem_autor = [p["nr"] for p in props if not p.get("parlamentar")]

    cab = f"""// GERADO AUTOMATICAMENTE a partir de Controle_LEXOR.xlsx — não editar à mão.
// Para atualizar: substitua a planilha na raiz do repositório e rode gerar_lexor.py.
//
// Propostas: {len(props)}   |   Ações cadastradas: {len(acoes)}
// Correções aplicadas na conversão:
//   · Ação em notação científica recuperada da Funcional Programática ({len([p for p in props if p.get("acao") == "2E74"])} casos de 2E74)
//   · Zeros à esquerda restaurados em Função / Subfunção / Programa
//   · UO lida da Funcional Programática (a tabela Ações é ambígua por ação)
//   · Objeto e Justificativa convertidos de CAIXA ALTA para caixa de frase
//   · Tipo de emenda padronizado em 3 categorias (+ marca de repetida)
//
// Pendências de preenchimento detectadas na origem:
//   · {len(sem_acao)} propostas sem ação correspondente na aba Ações
//   · {len(div_uo)} propostas com UO divergente entre a Funcional Programática e a aba Ações
//   · {len(sem_valor)} propostas sem valor em GND 3 e GND 4
//   · {len(negociados)} propostas com valor negociado (prevalece sobre o valor original)
//   · {len(sem_autor)} propostas sem parlamentar autor definido
"""

    js = cab + "\nexport const ACOES_LEXOR = " + json.dumps(acoes, ensure_ascii=False, indent=0) + ";\n"
    js += "\nexport const UO_LEXOR = " + json.dumps(uos, ensure_ascii=False, indent=0) + ";\n"
    js += "\nexport const PROPOSTAS_LEXOR = " + json.dumps(props, ensure_ascii=False, separators=(",", ":")) + ";\n"

    with open(DESTINO, "w", encoding="utf-8") as f:
        f.write(js)

    print(f"OK -> {DESTINO}")
    print(f"  propostas: {len(props)} | ações: {len(acoes)} | UOs: {len(uos)}")
    print(f"  sem ação na tabela: {len(sem_acao)}")
    print(f"  UO divergente:      {len(div_uo)}")
    print(f"  sem valor:          {len(sem_valor)}")
    print(f"  sem autor:          {len(sem_autor)}")


if __name__ == "__main__":
    main()
